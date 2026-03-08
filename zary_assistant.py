"""
ZARY & CO — Retail Bot v3.5 (FULL FILE)
✅ aiogram 3.x
✅ SQLite (bot.db)
✅ Admins only (ADMIN_ID_1..3)
✅ Channel notifications (CHANNEL_ID)
✅ Orders + Cart + Admin panel
✅ Excel export (manual)
✅ Render HTTP endpoints for Cron:
   - /cron/monthly?secret=...
   - /cron/daily?secret=...

✅ Weekly scheduled posts (Mon–Sat 18:00 Tashkent via cron)
✅ Sunday reminder to admin to upload new weekly posts

✅ ADDED (Telegram + Web analytics panel):
- Web admin dashboard: /admin?token=...
- Orders page + filters: /admin/orders?token=...
- API endpoints for charts + status update
- Funnel events + conversion: cart_add -> order_created
- /find <phone_part> for admins (quick CRM search)

✅ Improvements:
- "📞 Написать клиенту" button in admin message (tg://user?id=USER_ID)
- Different thank-you text after DELIVERED
- Fixed /admin/orders API URL bug
- Removed APScheduler dependency (asyncio loop)
- SQLite WAL + timeout to reduce delays/locks
"""

import os
import html
import asyncio
import json
from datetime import datetime, timedelta
from calendar import monthrange
from typing import Optional, Dict, List, Tuple
from pathlib import Path
import sqlite3
import threading

from zoneinfo import ZoneInfo
TZ = ZoneInfo("Asia/Tashkent")


# =========================
# ENV
# =========================
BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
if not BOT_TOKEN:
    raise RuntimeError("❌ BOT_TOKEN не установлен!")

ADMIN_IDS: List[int] = []
for i in range(1, 4):
    v = os.getenv(f"ADMIN_ID_{i}", "").strip()
    if v and v.lstrip("-").isdigit():
        ADMIN_IDS.append(int(v))

# fallback compatibility
if not ADMIN_IDS:
    old_admin = os.getenv("MANAGER_CHAT_ID", "").strip()
    if old_admin and old_admin.lstrip("-").isdigit():
        ADMIN_IDS.append(int(old_admin))

if not ADMIN_IDS:
    raise RuntimeError("❌ Нужен хотя бы один ADMIN_ID_1 (личный Telegram ID)")

PRIMARY_ADMIN = ADMIN_IDS[0]

_channel_id = os.getenv("CHANNEL_ID", "").strip()
CHANNEL_ID = int(_channel_id) if _channel_id and _channel_id.lstrip("-").isdigit() else None

CHANNEL_USERNAME = os.getenv("CHANNEL_USERNAME", "zaryco_official").strip().lstrip("@")
TG_CHANNEL_URL = f"https://t.me/{CHANNEL_USERNAME}"

PHONE = os.getenv("MANAGER_PHONE", "+998771202255").strip()
MANAGER_USERNAME = os.getenv("MANAGER_USERNAME", "zaryco_official").strip().lstrip("@")

PORT = int(os.getenv("PORT", "10000"))
DB_PATH = os.getenv("DB_PATH", "bot.db")

CRON_SECRET = os.getenv("CRON_SECRET", "").strip()
ADMIN_PANEL_TOKEN = os.getenv("ADMIN_PANEL_TOKEN", "").strip()

if not ADMIN_PANEL_TOKEN:
    print("⚠️ ADMIN_PANEL_TOKEN не установлен! /admin будет не защищен. Рекомендуется установить.")

# Follow links
FOLLOW_TG = "https://t.me/zaryco_official"
FOLLOW_YT = "https://www.youtube.com/@ZARYCOOFFICIAL"
FOLLOW_IG = "https://www.instagram.com/zary.co/"


# =========================
# PRODUCTS (Quick order list)
# =========================
PRODUCTS_RU = [
    "Худи детское", "Свитшот", "Футболка", "Рубашка", "Джинсы",
    "Брюки классические", "Юбка", "Платье", "Куртка демисезонная",
    "Костюм спортивный", "Школьная форма (комплект)", "Жилет школьный",
    "Кардиган", "Пижама", "Комплект (кофта+брюки)"
]
PRODUCTS_UZ = [
    "Bolalar hudi", "Sviter", "Futbolka", "Ko'ylak", "Jinsi",
    "Klassik shim", "Yubka", "Ko'ylak (dress)", "Demisezon kurtka",
    "Sport kostyum", "Maktab formasi (komplekt)", "Maktab jileti",
    "Kardigan", "Pijama", "Komplekt (kofta+shim)"
]


# =========================
# HELPERS
# =========================
def now_tz() -> datetime:
    return datetime.now(TZ)

def esc(s: str) -> str:
    return html.escape(str(s) if s else "")

def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS

def size_by_age(age: int) -> str:
    mapping = {1: "86", 2: "92", 3: "98", 4: "104", 5: "110", 6: "116",
               7: "122", 8: "128", 9: "134", 10: "140", 11: "146",
               12: "152", 13: "158", 14: "164", 15: "164"}
    return mapping.get(age, "122-128")

def size_by_height(height: int) -> str:
    sizes = [86, 92, 98, 104, 110, 116, 122, 128, 134, 140, 146, 152, 158, 164]
    closest = min(sizes, key=lambda x: abs(x - height))
    return str(closest)

def prev_month(dt: datetime) -> tuple[int, int]:
    first = dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    prev_last = first - timedelta(days=1)
    return prev_last.year, prev_last.month

def cron_allowed(secret: str) -> bool:
    return bool(CRON_SECRET) and secret == CRON_SECRET

def admin_panel_allowed(token: str) -> bool:
    if not ADMIN_PANEL_TOKEN:
        return True
    return token == ADMIN_PANEL_TOKEN


# =========================
# DB
# =========================
class Database:
    def __init__(self):
        self.db_path = DB_PATH
        self._local = threading.local()
        self._init_db()

    def _connect(self):
        conn = sqlite3.connect(self.db_path, check_same_thread=False, timeout=10)
        conn.row_factory = sqlite3.Row
        # Reduce locks + improve concurrency
        try:
            conn.execute("PRAGMA journal_mode=WAL;")
            conn.execute("PRAGMA synchronous=NORMAL;")
            conn.execute("PRAGMA temp_store=MEMORY;")
        except Exception:
            pass
        return conn

    def _get_conn(self):
        if not hasattr(self._local, "conn") or self._local.conn is None:
            self._local.conn = self._connect()
        return self._local.conn

    def _init_db(self):
        conn = self._connect()
        cur = conn.cursor()
        cur.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                username TEXT,
                lang TEXT DEFAULT 'ru',
                created_at TEXT
            );

            CREATE TABLE IF NOT EXISTS carts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                product_name TEXT,
                qty INTEGER DEFAULT 1,
                size TEXT,
                added_at TEXT DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                username TEXT,
                name TEXT,
                phone TEXT,
                city TEXT,
                items TEXT,
                total_amount INTEGER DEFAULT 0,
                delivery_type TEXT,
                delivery_address TEXT,
                comment TEXT,
                status TEXT DEFAULT 'new',
                manager_seen INTEGER DEFAULT 0,
                manager_id INTEGER,
                created_at TEXT,
                reminded_at TEXT
            );

            CREATE TABLE IF NOT EXISTS monthly_reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                year INTEGER,
                month INTEGER,
                sent_at TEXT,
                filename TEXT,
                total_orders INTEGER,
                total_amount INTEGER,
                status TEXT DEFAULT 'pending'
            );

            CREATE TABLE IF NOT EXISTS scheduled_posts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dow INTEGER,
                media_type TEXT,
                file_id TEXT,
                caption TEXT,
                week_key TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                posted_at TEXT
            );

            CREATE TABLE IF NOT EXISTS events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                event_type TEXT,
                meta TEXT,
                created_at TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_orders_user ON orders(user_id);
            CREATE INDEX IF NOT EXISTS idx_orders_status ON orders(status);
            CREATE INDEX IF NOT EXISTS idx_carts_user ON carts(user_id);
            CREATE INDEX IF NOT EXISTS idx_sched_week_dow ON scheduled_posts(week_key, dow);
            CREATE INDEX IF NOT EXISTS idx_events_type_time ON events(event_type, created_at);
        """)
        conn.commit()
        conn.close()

    # --- events
    def event_add(self, user_id: int, event_type: str, meta: Optional[Dict] = None):
        conn = self._get_conn()
        cur = conn.cursor()
        ts = now_tz().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute(
            "INSERT INTO events (user_id, event_type, meta, created_at) VALUES (?,?,?,?)",
            (user_id, event_type, json.dumps(meta or {}, ensure_ascii=False), ts)
        )
        conn.commit()

    # --- users
    def user_upsert(self, user_id: int, username: str, lang: str):
        conn = self._get_conn()
        cur = conn.cursor()
        ts = now_tz().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute("SELECT 1 FROM users WHERE user_id=?", (user_id,))
        if cur.fetchone():
            cur.execute("UPDATE users SET username=?, lang=? WHERE user_id=?", (username, lang, user_id))
        else:
            cur.execute(
                "INSERT INTO users (user_id, username, lang, created_at) VALUES (?,?,?,?)",
                (user_id, username, lang, ts)
            )
        conn.commit()

    def user_get(self, user_id: int) -> Optional[Dict]:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE user_id=?", (user_id,))
        row = cur.fetchone()
        return dict(row) if row else None

    # --- cart
    def cart_add(self, user_id: int, product_name: str, qty: int = 1, size: str = ""):
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("INSERT INTO carts (user_id, product_name, qty, size) VALUES (?,?,?,?)",
                    (user_id, product_name, qty, size))
        conn.commit()
        self.event_add(user_id, "cart_add", {"product": product_name, "qty": qty})

    def cart_get(self, user_id: int) -> List[Dict]:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("SELECT * FROM carts WHERE user_id=? ORDER BY id DESC", (user_id,))
        return [dict(r) for r in cur.fetchall()]

    def cart_clear(self, user_id: int):
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("DELETE FROM carts WHERE user_id=?", (user_id,))
        conn.commit()

    def cart_remove(self, cart_id: int):
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("DELETE FROM carts WHERE id=?", (cart_id,))
        conn.commit()

    # --- orders
    def order_create(self, data: Dict) -> int:
        conn = self._get_conn()
        cur = conn.cursor()
        ts = now_tz().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute("""
            INSERT INTO orders (
                user_id, username, name, phone, city, items,
                total_amount, delivery_type, delivery_address,
                comment, status, created_at
            )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            data["user_id"], data.get("username", ""), data["name"],
            data["phone"], data["city"], data["items"],
            data.get("total_amount", 0),
            data.get("delivery_type", ""),
            data.get("delivery_address", ""),
            data.get("comment", "—"),
            "new",
            ts
        ))
        conn.commit()
        order_id = cur.lastrowid
        self.event_add(data["user_id"], "order_created", {"order_id": order_id})
        return order_id

    def order_get(self, order_id: int) -> Optional[Dict]:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("SELECT * FROM orders WHERE id=?", (order_id,))
        row = cur.fetchone()
        return dict(row) if row else None

    def orders_get_by_status(self, status: str, limit: int = 50) -> List[Dict]:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("SELECT * FROM orders WHERE status=? ORDER BY created_at DESC LIMIT ?",
                    (status, limit))
        return [dict(r) for r in cur.fetchall()]

    def orders_get_user(self, user_id: int, limit: int = 10) -> List[Dict]:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("SELECT * FROM orders WHERE user_id=? ORDER BY id DESC LIMIT ?",
                    (user_id, limit))
        return [dict(r) for r in cur.fetchall()]

    def order_update_status(self, order_id: int, status: str, manager_id: int = None):
        conn = self._get_conn()
        cur = conn.cursor()
        if manager_id is not None:
            cur.execute("UPDATE orders SET status=?, manager_id=?, manager_seen=1 WHERE id=?",
                        (status, manager_id, order_id))
        else:
            cur.execute("UPDATE orders SET status=? WHERE id=?", (status, order_id))
        conn.commit()

    def order_mark_seen(self, order_id: int, manager_id: int):
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("UPDATE orders SET manager_seen=1, manager_id=? WHERE id=?",
                    (manager_id, order_id))
        conn.commit()

    def orders_get_for_reminder(self) -> List[Dict]:
        conn = self._get_conn()
        cur = conn.cursor()
        cutoff = (now_tz() - timedelta(minutes=30)).strftime("%Y-%m-%d %H:%M:%S")
        cur.execute("""
            SELECT * FROM orders
            WHERE status='new' AND manager_seen=0
              AND created_at < ?
              AND (reminded_at IS NULL OR reminded_at < ?)
            ORDER BY created_at DESC
        """, (cutoff, cutoff))
        return [dict(r) for r in cur.fetchall()]

    def order_update_reminded(self, order_id: int):
        conn = self._get_conn()
        cur = conn.cursor()
        ts = now_tz().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute("UPDATE orders SET reminded_at=? WHERE id=?", (ts, order_id))
        conn.commit()

    def orders_get_monthly(self, year: int, month: int) -> List[Dict]:
        conn = self._get_conn()
        cur = conn.cursor()
        start = f"{year}-{month:02d}-01 00:00:00"
        last_day = monthrange(year, month)[1]
        end = f"{year}-{month:02d}-{last_day} 23:59:59"
        cur.execute("SELECT * FROM orders WHERE created_at BETWEEN ? AND ? ORDER BY id",
                    (start, end))
        return [dict(r) for r in cur.fetchall()]

    def report_mark_sent(self, year: int, month: int, filename: str, total_orders: int, total_amount: int):
        conn = self._get_conn()
        cur = conn.cursor()
        ts = now_tz().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute("""
            INSERT INTO monthly_reports (year, month, sent_at, filename, total_orders, total_amount, status)
            VALUES (?,?,?,?,?,?,?)
        """, (year, month, ts, filename, total_orders, total_amount, "sent"))
        conn.commit()

    def report_is_sent(self, year: int, month: int) -> bool:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM monthly_reports WHERE year=? AND month=? AND status='sent'",
                    (year, month))
        return cur.fetchone() is not None

    def get_stats_all(self) -> Dict:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN status='new' THEN 1 ELSE 0 END) as new,
                SUM(CASE WHEN status='processing' THEN 1 ELSE 0 END) as processing,
                SUM(CASE WHEN status='shipped' THEN 1 ELSE 0 END) as shipped,
                SUM(CASE WHEN status='delivered' THEN 1 ELSE 0 END) as delivered,
                SUM(CASE WHEN status='cancelled' THEN 1 ELSE 0 END) as cancelled,
                COUNT(DISTINCT user_id) as unique_users
            FROM orders
        """)
        row = cur.fetchone()
        return dict(row) if row else {"total": 0, "new": 0, "processing": 0, "shipped": 0, "delivered": 0, "cancelled": 0, "unique_users": 0}

    # --- weekly scheduled posts
    def week_key_now(self, dt: datetime) -> str:
        iso = dt.isocalendar()
        return f"{iso[0]}-W{iso[1]:02d}"

    def sched_add(self, dow: int, media_type: str, file_id: str, caption: str, week_key: str):
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO scheduled_posts (dow, media_type, file_id, caption, week_key)
            VALUES (?,?,?,?,?)
        """, (dow, media_type, file_id, caption, week_key))
        conn.commit()

    def sched_get_for_day(self, dow: int, week_key: str) -> Optional[Dict]:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT * FROM scheduled_posts
            WHERE dow=? AND week_key=? AND posted_at IS NULL
            ORDER BY id ASC
            LIMIT 1
        """, (dow, week_key))
        row = cur.fetchone()
        return dict(row) if row else None

    def sched_mark_posted(self, post_id: int):
        conn = self._get_conn()
        cur = conn.cursor()
        ts = now_tz().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute("UPDATE scheduled_posts SET posted_at=? WHERE id=?", (ts, post_id))
        conn.commit()

    def sched_count_week(self, week_key: str) -> int:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) as c FROM scheduled_posts WHERE week_key=?", (week_key,))
        r = cur.fetchone()
        return int(r["c"]) if r else 0

    # --- web analytics helpers
    def stats_range(self, start: str, end: str) -> Dict:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN status='new' THEN 1 ELSE 0 END) as new,
                SUM(CASE WHEN status='processing' THEN 1 ELSE 0 END) as processing,
                SUM(CASE WHEN status='shipped' THEN 1 ELSE 0 END) as shipped,
                SUM(CASE WHEN status='delivered' THEN 1 ELSE 0 END) as delivered,
                SUM(CASE WHEN status='cancelled' THEN 1 ELSE 0 END) as cancelled
            FROM orders
            WHERE created_at BETWEEN ? AND ?
        """, (start, end))
        row = cur.fetchone()
        return dict(row) if row else {"total": 0, "new": 0, "processing": 0, "shipped": 0, "delivered": 0, "cancelled": 0}

    def top_products_range(self, start: str, end: str, limit: int = 10) -> List[Tuple[str, int]]:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("SELECT items FROM orders WHERE created_at BETWEEN ? AND ?", (start, end))
        counter: Dict[str, int] = {}
        for r in cur.fetchall():
            try:
                items = json.loads(r["items"] or "[]")
            except Exception:
                items = []
            for it in items:
                name = (it.get("name") or "").strip()
                qty = int(it.get("qty") or 1)
                if not name:
                    continue
                counter[name] = counter.get(name, 0) + qty
        return sorted(counter.items(), key=lambda x: x[1], reverse=True)[:limit]

    def top_cities_range(self, start: str, end: str, limit: int = 10) -> List[Tuple[str, int]]:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT city, COUNT(*) as c
            FROM orders
            WHERE created_at BETWEEN ? AND ?
            GROUP BY city
            ORDER BY c DESC
            LIMIT ?
        """, (start, end, limit))
        return [(r["city"] or "—", int(r["c"])) for r in cur.fetchall()]

    def ru_vs_uz_range(self, start: str, end: str) -> Dict:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT u.lang as lang, COUNT(o.id) as c
            FROM orders o
            LEFT JOIN users u ON u.user_id = o.user_id
            WHERE o.created_at BETWEEN ? AND ?
            GROUP BY u.lang
        """, (start, end))
        res = {"ru": 0, "uz": 0, "unknown": 0}
        for r in cur.fetchall():
            lang = (r["lang"] or "unknown").lower()
            if lang not in res:
                lang = "unknown"
            res[lang] += int(r["c"])
        return res

    def funnel_range(self, start: str, end: str) -> Dict:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT
              SUM(CASE WHEN event_type='cart_add' THEN 1 ELSE 0 END) as cart_add,
              SUM(CASE WHEN event_type='order_created' THEN 1 ELSE 0 END) as order_created
            FROM events
            WHERE created_at BETWEEN ? AND ?
        """, (start, end))
        r = cur.fetchone()
        cart_add = int(r["cart_add"] or 0) if r else 0
        order_created = int(r["order_created"] or 0) if r else 0
        conv = (order_created / cart_add * 100.0) if cart_add > 0 else 0.0
        return {"cart_add": cart_add, "order_created": order_created, "conversion": round(conv, 2)}

    def orders_filter(self, status: str = "", city: str = "", phone_q: str = "", limit: int = 200) -> List[Dict]:
        conn = self._get_conn()
        cur = conn.cursor()
        q = "SELECT * FROM orders WHERE 1=1"
        args: List = []
        if status:
            q += " AND status=?"
            args.append(status)
        if city:
            q += " AND city LIKE ?"
            args.append(f"%{city}%")
        if phone_q:
            q += " AND phone LIKE ?"
            args.append(f"%{phone_q}%")
        q += " ORDER BY id DESC LIMIT ?"
        args.append(limit)
        cur.execute(q, tuple(args))
        return [dict(r) for r in cur.fetchall()]

    def find_orders_by_phone(self, phone_part: str, limit: int = 20) -> List[Dict]:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT * FROM orders
            WHERE phone LIKE ?
            ORDER BY id DESC
            LIMIT ?
        """, (f"%{phone_part}%", limit))
        return [dict(r) for r in cur.fetchall()]


db = Database()


# =========================
# aiogram
# =========================
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message, CallbackQuery,
    InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup, KeyboardButton,
)
from aiogram.types.input_file import FSInputFile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


# =========================
# TEXTS
# =========================
TEXT = {
    "ru": {
        "welcome": "👋 Добро пожаловать в <b>ZARY & CO</b>!\n\n🧸 Детская одежда премиум качества\n📦 Доставка по Узбекистану 1-5 дней\n\nВыберите действие 👇",
        "menu": "📍 Главное меню",
        "catalog": "📸 <b>Каталог</b>\n\nВыберите категорию:",
        "catalog_hint": "Чтобы быстро оформить — нажмите <b>✅ Заказ</b> (выбор товаров внутри бота).",
        "price": "🧾 <b>Прайс-лист</b>\n\n💬 Цена — по договоренности (зависит от модели/размера).\n\n✅ Нажмите «Заказ» для оформления",
        "size": "📏 <b>Подбор размера</b>\n\nВыберите способ:",
        "size_age": "Введите возраст (1-15 лет):\nПример: 7",
        "size_height": "Введите рост в см:\nПример: 125",
        "size_result": "📏 Рекомендуемый размер: <b>{size}</b>\n\n✅ Если определились — нажмите <b>✅ Заказ</b> или вернитесь в меню.",
        "cart": "🛒 <b>Корзина</b>\n\n{items}\n\n💬 Цена: <b>по договоренности</b>\nНажмите <b>✅ Оформить</b>, чтобы продолжить.",
        "cart_empty": "🛒 Корзина пуста\n\nПерейдите в <b>✅ Заказ</b> и выберите товар (или напишите свой).",
        "cart_added": "✅ Добавлено в корзину",
        "delivery": "🚚 <b>Доставка</b>\n\n1️⃣ <b>B2B Почта</b> — 2-5 дней, весь Узбекистан\n2️⃣ <b>Яндекс Курьер</b> — 1-3 дня, крупные города\n3️⃣ <b>Яндекс ПВЗ</b> — 1-3 дня, пункты выдачи\n\n💬 Стоимость доставки зависит от города.",
        "faq": "❓ <b>FAQ</b>\n\n<b>Доставка?</b>\n— По всему Узбекистану, 1-5 дней\n\n<b>Оплата?</b>\n— Наличными или переводом\n\n<b>Возврат?</b>\n— 14 дней при сохранении вида\n\n<b>Размеры?</b>\n— Используйте подбор в боте",
        "contact": "📞 <b>Связаться</b>\n\n☎️ {phone}\n⏰ Пн-Пт: 09:00-21:00\n📱 @{username}",
        "order_start": "🛍 <b>Выберите товар</b>\n\nНажмите на кнопку товара ниже 👇\nЕсли вашего товара нет — нажмите <b>✍️ Ввести вручную</b>",
        "order_manual": "📝 Введите название товара (например: худи, джинсы, школьная форма):",
        "order_phone": "📱 Отправьте номер телефона:",
        "order_city": "🏙 Введите город:",
        "order_delivery": "🚚 Выберите способ доставки (нажмите кнопку):",
        "order_address": "📍 Введите адрес доставки:",
        "order_confirm": "📝 <b>Проверьте заказ:</b>\n\n👤 {name}\n📱 {phone}\n🏙 {city}\n🚚 {delivery}\n📍 {address}\n\n🛒 Товары:\n{items}\n\n💬 Цена: <b>по договоренности</b>\nМенеджер уточнит размер и итоговую сумму.\n\nПодтвердить?",
        "order_success": "✅ Заказ #{order_id} принят!\n\nУважаемый покупатель, вам поступят уведомления о статусе.\nМенеджер скоро свяжется и уточнит детали.\n⏰ 09:00-21:00",

        # ✅ оставляем как было — для "принят"
        "thanks_new": "🙏 Спасибо за заказ! Мы рады, что вы с нами 🤍\n\nЧтобы нас не потерять — подпишитесь на наши каналы:",

        # ✅ новый текст — для "доставлен"
        "thanks_delivered": (
            "🤍 Спасибо, что выбрали ZARY & CO!\n\n"
            "Надеемся, одежда принесёт радость и комфорт.\n"
            "Носите с удовольствием и на здоровье ✨\n\n"
            "Будем рады видеть вас снова!\n"
            "Чтобы не пропустить новинки — подпишитесь на наши каналы 👇"
        ),

        "history": "📜 <b>История заказов</b>\n\n{orders}",
        "history_empty": "📜 У вас пока нет заказов",
        "admin_menu": "🛠 <b>Админ панель</b>\n\nВыберите действие:",
        "admin_stats": "📊 <b>Статистика</b>\n\n📦 Всего: {total}\n🆕 Новых: {new}\n⚙️ В обработке: {processing}\n🚚 Отправлено: {shipped}\n✅ Доставлено: {delivered}\n❌ Отменено: {cancelled}\n👥 Клиентов: {unique_users}",
        "cancelled": "❌ Отменено",
    },
    "uz": {
        "welcome": "👋 <b>ZARY & CO</b> ga xush kelibsiz!\n\n🧸 Bolalar kiyimi premium sifat\n📦 O'zbekiston bo'ylab yetkazib berish 1-5 kun\n\nAmalni tanlang 👇",
        "menu": "📍 Asosiy menyu",
        "catalog": "📸 <b>Katalog</b>\n\nKategoriyani tanlang:",
        "catalog_hint": "Tez buyurtma uchun <b>✅ Buyurtma</b> ni bosing (bot ichida tanlash).",
        "price": "🧾 <b>Narxlar</b>\n\n💬 Narx — kelishuv bo'yicha (model/o'lchamga qarab).\n\n✅ «Buyurtma» ni bosing",
        "size": "📏 <b>O'lcham tanlash</b>\n\nUsulni tanlang:",
        "size_age": "Yoshini kiriting (1-15 yosh):\nMisol: 7",
        "size_height": "Bo'yni sm da kiriting:\nMisol: 125",
        "size_result": "📏 Tavsiya etilgan o'lcham: <b>{size}</b>\n\n✅ Tayyor bo'lsangiz <b>✅ Buyurtma</b> ni bosing yoki menyuga qayting.",
        "cart": "🛒 <b>Savat</b>\n\n{items}\n\n💬 Narx: <b>kelishuv bo'yicha</b>\n<b>✅ Rasmiylashtirish</b> ni bosing.",
        "cart_empty": "🛒 Savat bo'sh\n\n<b>✅ Buyurtma</b> ga kiring va tovar tanlang (yoki o'zingiz yozing).",
        "cart_added": "✅ Savatga qo'shildi",
        "delivery": "🚚 <b>Yetkazib berish</b>\n\n1️⃣ <b>B2B Pochta</b> — 2-5 kun\n2️⃣ <b>Yandex Kuryer</b> — 1-3 kun\n3️⃣ <b>Yandex PVZ</b> — 1-3 kun\n\n💬 Yetkazib berish narxi shahar bo'yicha.",
        "faq": "❓ <b>FAQ</b>\n\n<b>Yetkazib berish?</b>\n— Butun O'zbekiston, 1-5 kun\n\n<b>To'lov?</b>\n— Naqd yoki o'tkazma\n\n<b>Qaytarish?</b>\n— 14 kun ichida (tovar ko'rinishi saqlangan bo'lsa)\n\n<b>O'lchamlar?</b>\n— Botdagi o'lcham tanlashdan foydalaning",
        "contact": "📞 <b>Aloqa</b>\n\n☎️ {phone}\n⏰ Du-Sha: 09:00-21:00\n📱 @{username}",
        "order_start": "🛍 <b>Tovar tanlang</b>\n\nQuyidagi tugmalardan birini bosing 👇\nAgar kerakli tovar bo'lmasa — <b>✍️ Qo'lda kiritish</b> ni bosing",
        "order_manual": "📝 Mahsulot nomini kiriting (masalan: hudi, jinsi, maktab formasi):",
        "order_phone": "📱 Telefon raqamingizni yuboring:",
        "order_city": "🏙 Shaharni kiriting:",
        "order_delivery": "🚚 Yetkazib berish usulini tanlang (tugmani bosing):",
        "order_address": "📍 Manzilni kiriting:",
        "order_confirm": "📝 <b>Buyurtmani tekshiring:</b>\n\n👤 {name}\n📱 {phone}\n🏙 {city}\n🚚 {delivery}\n📍 {address}\n\n🛒 Tovarlar:\n{items}\n\n💬 Narx: <b>kelishuv bo'yicha</b>\nMenejer o'lcham va yakuniy summani aniqlaydi.\n\nTasdiqlaysizmi?",
        "order_success": "✅ Buyurtma #{order_id} qabul qilindi!\n\nHurmatli mijoz, status bo'yicha xabarlar yuboriladi.\nMenejer tez orada bog'lanadi.\n⏰ 09:00-21:00",

        "thanks_new": "🙏 Buyurtmangiz uchun rahmat! Siz biz bilan ekaningizdan xursandmiz 🤍\n\nBizni yo‘qotib qo‘ymaslik uchun kanallarimizga obuna bo‘ling:",
        "thanks_delivered": (
            "🤍 ZARY & CO ni tanlaganingiz uchun rahmat!\n\n"
            "Kiyim sizga qulaylik va xursandchilik olib kelsin.\n"
            "Yaxshi kayfiyat bilan kiying ✨\n\n"
            "Yana sizni ko‘rishdan xursand bo‘lamiz!\n"
            "Yangiliklarni o‘tkazib yubormaslik uchun kanallarimizga obuna bo‘ling 👇"
        ),

        "history": "📜 <b>Buyurtmalar tarixi</b>\n\n{orders}",
        "history_empty": "📜 Hozircha buyurtmalar yo'q",
        "admin_menu": "🛠 <b>Admin paneli</b>\n\nAmalni tanlang:",
        "admin_stats": "📊 <b>Statistika</b>\n\n📦 Jami: {total}\n🆕 Yangi: {new}\n⚙️ Ishlanmoqda: {processing}\n🚚 Jo'natildi: {shipped}\n✅ Yetkazildi: {delivered}\n❌ Bekor: {cancelled}\n👥 Mijozlar: {unique_users}",
        "cancelled": "❌ Bekor qilindi",
    }
}


# =========================
# KEYBOARDS
# =========================
def kb_main(lang: str, is_admin_flag: bool = False) -> ReplyKeyboardMarkup:
    if lang == "uz":
        rows = [
            [KeyboardButton(text="📸 Katalog"), KeyboardButton(text="🧾 Narxlar")],
            [KeyboardButton(text="📏 O'lcham"), KeyboardButton(text="🛒 Savat")],
            [KeyboardButton(text="🚚 Yetkazib berish"), KeyboardButton(text="❓ FAQ")],
            [KeyboardButton(text="📞 Aloqa"), KeyboardButton(text="✅ Buyurtma")],
            [KeyboardButton(text="📜 Buyurtmalar"), KeyboardButton(text="🌐 Til")],
        ]
    else:
        rows = [
            [KeyboardButton(text="📸 Каталог"), KeyboardButton(text="🧾 Прайс")],
            [KeyboardButton(text="📏 Размер"), KeyboardButton(text="🛒 Корзина")],
            [KeyboardButton(text="🚚 Доставка"), KeyboardButton(text="❓ FAQ")],
            [KeyboardButton(text="📞 Связаться"), KeyboardButton(text="✅ Заказ")],
            [KeyboardButton(text="📜 История"), KeyboardButton(text="🌐 Язык")],
        ]
    if is_admin_flag:
        rows.append([KeyboardButton(text="🛠 Админ" if lang == "ru" else "🛠 Admin")])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)

def kb_catalog(lang: str) -> InlineKeyboardMarkup:
    cats = [
        [("👶 Мальчики", "cat:boys"), ("👧 Девочки", "cat:girls")],
        [("🧒 Унисекс", "cat:unisex"), ("🎒 Школа", "cat:school")],
        [("🔥 Новинки", "cat:new"), ("💰 Акции", "cat:sale")],
    ]
    buttons = []
    for row in cats:
        buttons.append([
            InlineKeyboardButton(text=row[0][0], callback_data=row[0][1]),
            InlineKeyboardButton(text=row[1][0], callback_data=row[1][1])
        ])
    buttons.append([InlineKeyboardButton(text="✅ Быстрый заказ" if lang=="ru" else "✅ Tez buyurtma", callback_data="quick_order")])
    buttons.append([InlineKeyboardButton(text="⬅️ Назад" if lang=="ru" else "⬅️ Orqaga", callback_data="back:menu")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def kb_size(lang: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="👶 По возрасту" if lang == "ru" else "👶 Yosh bo'yicha", callback_data="size:age")],
        [InlineKeyboardButton(text="📏 По росту" if lang == "ru" else "📏 Bo'y bo'yicha", callback_data="size:height")],
        [InlineKeyboardButton(text="⬅️ Назад" if lang == "ru" else "⬅️ Orqaga", callback_data="back:menu")],
    ])

def kb_delivery(lang: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📦 B2B Почта" if lang=="ru" else "📦 B2B Pochta", callback_data="delivery:b2b")],
        [InlineKeyboardButton(text="🚚 Яндекс Курьер" if lang=="ru" else "🚚 Yandex Kuryer", callback_data="delivery:yandex_courier")],
        [InlineKeyboardButton(text="🏪 Яндекс ПВЗ" if lang=="ru" else "🏪 Yandex PVZ", callback_data="delivery:yandex_pvz")],
        [InlineKeyboardButton(text="⬅️ Назад" if lang == "ru" else "⬅️ Orqaga", callback_data="back:menu")],
    ])

def kb_cart(items: List[Dict], lang: str) -> InlineKeyboardMarkup:
    buttons = []
    for item in items:
        name = item["product_name"][:22]
        btn_text = f"❌ {name} (x{item['qty']})"
        buttons.append([InlineKeyboardButton(text=btn_text, callback_data=f"cart_remove:{item['id']}")])

    buttons.extend([
        [InlineKeyboardButton(text="✅ Оформить" if lang == "ru" else "✅ Rasmiylashtirish", callback_data="cart:checkout")],
        [InlineKeyboardButton(text="🧹 Очистить" if lang == "ru" else "🧹 Tozalash", callback_data="cart:clear")],
        [InlineKeyboardButton(text="⬅️ Назад" if lang == "ru" else "⬅️ Orqaga", callback_data="back:menu")],
    ])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def kb_order_confirm(lang: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Подтвердить" if lang=="ru" else "✅ Tasdiqlash", callback_data="order:confirm")],
        [InlineKeyboardButton(text="❌ Отмена" if lang=="ru" else "❌ Bekor", callback_data="order:cancel")],
    ])

def kb_admin(lang: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📋 Новые заказы" if lang == "ru" else "📋 Yangi buyurtmalar", callback_data="admin:new")],
        [InlineKeyboardButton(text="⚙️ В обработке" if lang == "ru" else "⚙️ Ishlanmoqda", callback_data="admin:processing")],
        [InlineKeyboardButton(text="📊 Статистика" if lang == "ru" else "📊 Statistika", callback_data="admin:stats")],
        [InlineKeyboardButton(text="📤 Excel отчет" if lang == "ru" else "📤 Excel hisobot", callback_data="admin:export")],
        [InlineKeyboardButton(text="📰 Посты недели" if lang == "ru" else "📰 Haftalik postlar", callback_data="admin:posts")],
        [InlineKeyboardButton(text="⬅️ Назад" if lang == "ru" else "⬅️ Orqaga", callback_data="back:menu")],
    ])

# ✅ NEW: user_id inside admin keyboard (write to client)
def kb_admin_order(order_id: int, user_id: int, lang: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📞 Написать клиенту" if lang=="ru" else "📞 Mijozga yozish", url=f"tg://user?id={user_id}")],
        [
            InlineKeyboardButton(text="👁 Просмотрено" if lang=="ru" else "👁 Ko'rildi", callback_data=f"order_seen:{order_id}"),
            InlineKeyboardButton(text="⚙️ В работу" if lang=="ru" else "⚙️ Ishga", callback_data=f"order_process:{order_id}")
        ],
        [
            InlineKeyboardButton(text="🚚 Отправлен" if lang=="ru" else "🚚 Jo'natildi", callback_data=f"order_ship:{order_id}"),
            InlineKeyboardButton(text="✅ Доставлен" if lang=="ru" else "✅ Yetkazildi", callback_data=f"order_deliver:{order_id}")
        ],
        [InlineKeyboardButton(text="❌ Отмена" if lang=="ru" else "❌ Bekor", callback_data=f"order_cancel:{order_id}")],
    ])

def kb_contact(lang: str) -> ReplyKeyboardMarkup:
    if lang == "uz":
        btn = KeyboardButton(text="📱 Raqamni yuborish", request_contact=True)
        cancel = KeyboardButton(text="❌ Bekor qilish")
    else:
        btn = KeyboardButton(text="📱 Отправить номер", request_contact=True)
        cancel = KeyboardButton(text="❌ Отмена")
    return ReplyKeyboardMarkup(keyboard=[[btn], [cancel]], resize_keyboard=True, one_time_keyboard=True)

def kb_channel_and_menu(lang: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📣 Канал" if lang=="ru" else "📣 Kanal", url=TG_CHANNEL_URL)],
        [InlineKeyboardButton(text="⬅️ Меню" if lang=="ru" else "⬅️ Menyu", callback_data="back:menu")],
    ])

def kb_follow_links(lang: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📣 Telegram канал", url=FOLLOW_TG)],
        [InlineKeyboardButton(text="📺 YouTube", url=FOLLOW_YT)],
        [InlineKeyboardButton(text="📸 Instagram", url=FOLLOW_IG)],
    ])

def kb_quick_products(lang: str) -> InlineKeyboardMarkup:
    items = PRODUCTS_RU if lang == "ru" else PRODUCTS_UZ
    rows = []
    for i in range(0, min(len(items), 12), 2):
        a = items[i]
        b = items[i + 1] if i + 1 < min(len(items), 12) else None
        row = [InlineKeyboardButton(text=a, callback_data=f"prod:{i}")]
        if b:
            row.append(InlineKeyboardButton(text=b, callback_data=f"prod:{i+1}"))
        rows.append(row)

    rows.append([InlineKeyboardButton(text="✍️ Ввести вручную" if lang=="ru" else "✍️ Qo'lda kiritish", callback_data="prod_manual")])
    rows.append([InlineKeyboardButton(text="🛒 Корзина" if lang=="ru" else "🛒 Savat", callback_data="go_cart")])
    rows.append([InlineKeyboardButton(text="⬅️ Меню" if lang=="ru" else "⬅️ Menyu", callback_data="back:menu")])
    return InlineKeyboardMarkup(inline_keyboard=rows)

def kb_after_add(lang: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Добавить ещё" if lang=="ru" else "➕ Yana qo‘shish", callback_data="quick_order")],
        [InlineKeyboardButton(text="🛒 Перейти в корзину" if lang=="ru" else "🛒 Savatga o‘tish", callback_data="go_cart")],
        [InlineKeyboardButton(text="✅ Оформить заказ" if lang=="ru" else "✅ Buyurtmani rasmiylashtirish", callback_data="cart:checkout")],
        [InlineKeyboardButton(text="⬅️ Меню" if lang=="ru" else "⬅️ Menyu", callback_data="back:menu")],
    ])

def kb_dow(lang: str) -> InlineKeyboardMarkup:
    if lang == "uz":
        names = [(1, "Dushanba"), (2, "Seshanba"), (3, "Chorshanba"), (4, "Payshanba"), (5, "Juma"), (6, "Shanba")]
    else:
        names = [(1, "Понедельник"), (2, "Вторник"), (3, "Среда"), (4, "Четверг"), (5, "Пятница"), (6, "Суббота")]

    rows = []
    for i in range(0, 6, 2):
        a = names[i]
        b = names[i + 1]
        rows.append([
            InlineKeyboardButton(text=a[1], callback_data=f"dow:{a[0]}"),
            InlineKeyboardButton(text=b[1], callback_data=f"dow:{b[0]}")
        ])
    rows.append([InlineKeyboardButton(text="⬅️ Назад" if lang == "ru" else "⬅️ Orqaga", callback_data="admin:back")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


# =========================
# FSM
# =========================
class States(StatesGroup):
    size_age = State()
    size_height = State()

    order_name = State()
    order_phone = State()
    order_city = State()
    order_delivery = State()
    order_address = State()

    prod_manual = State()

    admin_post_dow = State()
    admin_post_media = State()


# =========================
# BOT INIT
# =========================
bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())


# =========================
# HANDLERS
# =========================
@dp.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()
    user_id = message.from_user.id
    username = message.from_user.username or ""
    lang = "uz" if (message.from_user.language_code == "uz") else "ru"
    db.user_upsert(user_id, username, lang)
    await message.answer(TEXT[lang]["welcome"], reply_markup=kb_main(lang, is_admin(user_id)))

@dp.message(F.text.in_(["🌐 Язык", "🌐 Til"]))
async def cmd_lang(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = "uz" if user and user.get("lang") == "ru" else "ru"
    db.user_upsert(message.from_user.id, message.from_user.username or "", lang)
    await message.answer(TEXT[lang]["welcome"], reply_markup=kb_main(lang, is_admin(message.from_user.id)))

@dp.callback_query(F.data == "back:menu")
async def back_menu(call: CallbackQuery, state: FSMContext):
    await state.clear()
    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"
    await call.message.answer(TEXT[lang]["menu"], reply_markup=kb_main(lang, is_admin(call.from_user.id)))
    await call.answer()

# Catalog
@dp.message(F.text.in_(["📸 Каталог", "📸 Katalog"]))
async def cmd_catalog(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    await message.answer(TEXT[lang]["catalog"], reply_markup=kb_catalog(lang))
    await message.answer(TEXT[lang]["catalog_hint"], reply_markup=kb_main(lang, is_admin(message.from_user.id)))

@dp.callback_query(F.data.startswith("cat:"))
async def cat_select(call: CallbackQuery, state: FSMContext):
    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"
    cat = call.data.split(":")[1]
    await call.message.answer(
        f"📸 {cat.upper()}\n\nСмотрите полный каталог в канале 👇" if lang == "ru"
        else f"📸 {cat.upper()}\n\nTo'liq katalog kanalimizda 👇",
        reply_markup=kb_channel_and_menu(lang)
    )
    await call.answer()

@dp.callback_query(F.data == "quick_order")
async def quick_order(call: CallbackQuery, state: FSMContext):
    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"
    await state.clear()
    await call.message.answer(TEXT[lang]["order_start"], reply_markup=kb_quick_products(lang))
    await call.answer()

# Price
@dp.message(F.text.in_(["🧾 Прайс", "🧾 Narxlar"]))
async def cmd_price(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    await message.answer(TEXT[lang]["price"], reply_markup=kb_main(lang, is_admin(message.from_user.id)))

# Size
@dp.message(F.text.in_(["📏 Размер", "📏 O'lcham"]))
async def cmd_size(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    await message.answer(TEXT[lang]["size"], reply_markup=kb_size(lang))

@dp.callback_query(F.data.startswith("size:"))
async def size_select(call: CallbackQuery, state: FSMContext):
    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"
    mode = call.data.split(":")[1]
    if mode == "age":
        await state.set_state(States.size_age)
        await call.message.answer(TEXT[lang]["size_age"])
    else:
        await state.set_state(States.size_height)
        await call.message.answer(TEXT[lang]["size_height"])
    await call.answer()

@dp.message(States.size_age)
async def size_age_input(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    if not message.text or not message.text.isdigit():
        await message.answer(TEXT[lang]["size_age"])
        return
    age = int(message.text)
    if not (1 <= age <= 15):
        await message.answer(TEXT[lang]["size_age"])
        return
    size = size_by_age(age)
    await message.answer(TEXT[lang]["size_result"].format(size=size), reply_markup=kb_main(lang, is_admin(message.from_user.id)))
    await state.clear()

@dp.message(States.size_height)
async def size_height_input(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    if not message.text or not message.text.isdigit():
        await message.answer(TEXT[lang]["size_height"])
        return
    height = int(message.text)
    if not (50 <= height <= 180):
        await message.answer(TEXT[lang]["size_height"])
        return
    size = size_by_height(height)
    await message.answer(TEXT[lang]["size_result"].format(size=size), reply_markup=kb_main(lang, is_admin(message.from_user.id)))
    await state.clear()

# FAQ
@dp.message(F.text.in_(["❓ FAQ"]))
async def cmd_faq(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    await message.answer(TEXT[lang]["faq"], reply_markup=kb_channel_and_menu(lang))
    await message.answer(TEXT[lang]["menu"], reply_markup=kb_main(lang, is_admin(message.from_user.id)))

# Delivery info
@dp.message(F.text.in_(["🚚 Доставка", "🚚 Yetkazib berish"]))
async def cmd_delivery(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    await message.answer(TEXT[lang]["delivery"], reply_markup=kb_delivery(lang))

# Contact
@dp.message(F.text.in_(["📞 Связаться", "📞 Aloqa"]))
async def cmd_contact(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    text = TEXT[lang]["contact"].format(phone=PHONE, username=MANAGER_USERNAME or CHANNEL_USERNAME)
    await message.answer(text, reply_markup=kb_main(lang, is_admin(message.from_user.id)))

# Cart
@dp.message(F.text.in_(["🛒 Корзина", "🛒 Savat"]))
async def cmd_cart(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"

    items = db.cart_get(message.from_user.id)
    if not items:
        await message.answer(TEXT[lang]["cart_empty"], reply_markup=kb_main(lang, is_admin(message.from_user.id)))
        return

    items_text = "\n".join([f"• {esc(it['product_name'])} x{it['qty']}" for it in items])
    text = TEXT[lang]["cart"].format(items=items_text)
    await message.answer(text, reply_markup=kb_cart(items, lang))

@dp.callback_query(F.data == "go_cart")
async def go_cart(call: CallbackQuery, state: FSMContext):
    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"
    items = db.cart_get(call.from_user.id)
    if not items:
        await call.message.answer(TEXT[lang]["cart_empty"], reply_markup=kb_main(lang, is_admin(call.from_user.id)))
    else:
        items_text = "\n".join([f"• {esc(it['product_name'])} x{it['qty']}" for it in items])
        text = TEXT[lang]["cart"].format(items=items_text)
        await call.message.answer(text, reply_markup=kb_cart(items, lang))
    await call.answer()

@dp.callback_query(F.data.startswith("cart_remove:"))
async def cart_remove(call: CallbackQuery, state: FSMContext):
    cart_id = int(call.data.split(":")[1])
    db.cart_remove(cart_id)

    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"

    items = db.cart_get(call.from_user.id)
    if not items:
        await call.message.edit_text(TEXT[lang]["cart_empty"])
    else:
        items_text = "\n".join([f"• {esc(it['product_name'])} x{it['qty']}" for it in items])
        text = TEXT[lang]["cart"].format(items=items_text)
        await call.message.edit_text(text, reply_markup=kb_cart(items, lang))

    await call.answer("❌ Удалено" if lang == "ru" else "❌ O'chirildi")

@dp.callback_query(F.data == "cart:clear")
async def cart_clear(call: CallbackQuery, state: FSMContext):
    db.cart_clear(call.from_user.id)
    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"
    await call.message.edit_text(TEXT[lang]["cart_empty"])
    await call.answer()

@dp.callback_query(F.data == "cart:checkout")
async def cart_checkout(call: CallbackQuery, state: FSMContext):
    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"

    items = db.cart_get(call.from_user.id)
    if not items:
        await call.answer("Корзина пуста!" if lang == "ru" else "Savat bo'sh!")
        return

    await state.set_state(States.order_name)
    await call.message.answer("Введите ваше имя:" if lang == "ru" else "Ismingizni kiriting:")
    await call.answer()

# Quick products select
@dp.callback_query(F.data.startswith("prod:"))
async def prod_select(call: CallbackQuery, state: FSMContext):
    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"
    idx = int(call.data.split(":")[1])

    items = PRODUCTS_RU if lang == "ru" else PRODUCTS_UZ
    if 0 <= idx < len(items):
        db.cart_add(call.from_user.id, items[idx], 1)

        await call.message.answer(TEXT[lang]["cart_added"])
        await call.message.answer(
            ("🛒 Товар добавлен!\n\n"
             "Если хотите — добавьте ещё товары.\n"
             "Если достаточно — перейдите в корзину и оформите заказ 👇")
            if lang == "ru" else
            ("🛒 Mahsulot savatga qo‘shildi!\n\n"
             "Xohlasangiz yana qo‘shing.\n"
             "Yetarli bo‘lsa savatga o‘ting va buyurtmani rasmiylashtiring 👇"),
            reply_markup=kb_after_add(lang)
        )

    await call.answer()

@dp.callback_query(F.data == "prod_manual")
async def prod_manual_start(call: CallbackQuery, state: FSMContext):
    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"
    await state.set_state(States.prod_manual)
    await call.message.answer(TEXT[lang]["order_manual"])
    await call.answer()

@dp.message(States.prod_manual)
async def prod_manual_input(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    if not message.text or len(message.text.strip()) < 2:
        await message.answer(TEXT[lang]["order_manual"])
        return
    db.cart_add(message.from_user.id, message.text.strip(), 1)
    await message.answer(TEXT[lang]["cart_added"], reply_markup=kb_main(lang, is_admin(message.from_user.id)))
    await state.clear()

# Order flow entry
@dp.message(F.text.in_(["✅ Заказ", "✅ Buyurtma"]))
async def cmd_order(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    await state.clear()
    await message.answer(TEXT[lang]["order_start"], reply_markup=kb_quick_products(lang))

# Admin CRM quick find
@dp.message(F.text.startswith("/find"))
async def admin_find(message: Message):
    if not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2 or len(parts[1].strip()) < 3:
        await message.answer("Пример: /find 99877\nИщу по номеру телефона (часть номера).")
        return
    q = parts[1].strip()
    rows = db.find_orders_by_phone(q, limit=20)
    if not rows:
        await message.answer(f"Ничего не найдено по: {esc(q)}")
        return
    lines = []
    for o in rows[:10]:
        lines.append(f"#{o['id']} • {o['created_at'][:16]} • {esc(o['name'])} • {esc(o['phone'])} • {esc(o['city'])} • {o['status']}")
    await message.answer("🔎 Найдено:\n" + "\n".join(lines))

# Order steps
@dp.message(States.order_name)
async def order_name(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    if not message.text or len(message.text.strip()) < 2:
        await message.answer("Введите ваше имя:" if lang == "ru" else "Ismingizni kiriting:")
        return
    await state.update_data(name=message.text.strip())
    await state.set_state(States.order_phone)
    await message.answer(TEXT[lang]["order_phone"], reply_markup=kb_contact(lang))

@dp.message(States.order_phone)
async def order_phone(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    phone = message.contact.phone_number if message.contact else (message.text or "").strip()
    if not phone:
        await message.answer(TEXT[lang]["order_phone"], reply_markup=kb_contact(lang))
        return
    await state.update_data(phone=phone)
    await state.set_state(States.order_city)
    await message.answer(TEXT[lang]["order_city"], reply_markup=kb_main(lang, is_admin(message.from_user.id)))

@dp.message(States.order_city)
async def order_city(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    if not message.text or len(message.text.strip()) < 2:
        await message.answer(TEXT[lang]["order_city"])
        return
    await state.update_data(city=message.text.strip())
    await state.set_state(States.order_delivery)
    await message.answer(TEXT[lang]["order_delivery"], reply_markup=kb_delivery(lang))

@dp.message(States.order_delivery)
async def order_delivery_text_guard(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    await message.answer(TEXT[lang]["order_delivery"], reply_markup=kb_delivery(lang))

@dp.callback_query(F.data.startswith("delivery:"))
async def order_delivery_callback(call: CallbackQuery, state: FSMContext):
    delivery_type = call.data.split(":")[1]
    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"

    delivery_names = {
        "b2b": "B2B Почта" if lang == "ru" else "B2B Pochta",
        "yandex_courier": "Яндекс Курьер" if lang == "ru" else "Yandex Kuryer",
        "yandex_pvz": "Яндекс ПВЗ" if lang == "ru" else "Yandex PVZ"
    }

    await state.update_data(delivery=delivery_type, delivery_name=delivery_names.get(delivery_type, delivery_type))
    await state.set_state(States.order_address)
    await call.message.answer(TEXT[lang]["order_address"])
    await call.answer()

@dp.message(States.order_address)
async def order_address(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    if not message.text or len(message.text.strip()) < 3:
        await message.answer(TEXT[lang]["order_address"])
        return

    await state.update_data(address=message.text.strip())

    data = await state.get_data()
    items = db.cart_get(message.from_user.id)
    if not items:
        await state.clear()
        await message.answer(TEXT[lang]["cart_empty"], reply_markup=kb_main(lang, is_admin(message.from_user.id)))
        return

    items_text = "\n".join([f"• {esc(it['product_name'])} x{it['qty']}" for it in items])

    text = TEXT[lang]["order_confirm"].format(
        name=esc(data["name"]),
        phone=esc(data["phone"]),
        city=esc(data["city"]),
        delivery=esc(data.get("delivery_name", "—")),
        address=esc(data["address"]),
        items=items_text
    )
    await message.answer(text, reply_markup=kb_order_confirm(lang))

@dp.callback_query(F.data == "order:confirm")
async def order_confirm(call: CallbackQuery, state: FSMContext):
    user_row = db.user_get(call.from_user.id)
    lang = user_row["lang"] if user_row else "ru"
    data = await state.get_data()

    items = db.cart_get(call.from_user.id)
    if not items:
        await state.clear()
        await call.message.answer(TEXT[lang]["cart_empty"], reply_markup=kb_main(lang, is_admin(call.from_user.id)))
        await call.answer()
        return

    items_json = json.dumps([{"name": it["product_name"], "qty": it["qty"]} for it in items], ensure_ascii=False)

    order_data = {
        "user_id": call.from_user.id,
        "username": call.from_user.username or "",
        "name": data.get("name", "—"),
        "phone": data.get("phone", "—"),
        "city": data.get("city", "—"),
        "items": items_json,
        "total_amount": 0,
        "delivery_type": data.get("delivery", ""),
        "delivery_address": data.get("address", ""),
        "comment": "—",
    }

    order_id = db.order_create(order_data)

    # notify admins
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(
                admin_id,
                f"🆕 Новый заказ #{order_id}\n\n"
                f"👤 {esc(order_data['name'])}\n"
                f"📱 {esc(order_data['phone'])}\n"
                f"🏙 {esc(order_data['city'])}\n"
                f"🚚 {esc(data.get('delivery_name','—'))}\n"
                f"📍 {esc(order_data['delivery_address'])}\n"
                f"🛒 {', '.join([esc(it['product_name']) for it in items])}\n"
                f"💬 Цена: по договоренности",
                reply_markup=kb_admin_order(order_id, order_data["user_id"], "ru")
            )
        except Exception as e:
            print(f"Failed to notify admin {admin_id}: {e}")

    # notify channel (optional)
    if CHANNEL_ID:
        try:
            await bot.send_message(
                CHANNEL_ID,
                f"🆕 Новый заказ #{order_id}\n"
                f"👤 {esc(order_data['name'])}\n"
                f"📱 {esc(order_data['phone'])}\n"
                f"🏙 {esc(order_data['city'])}\n"
                f"🛒 {', '.join([esc(it['product_name']) for it in items])}\n"
                f"💬 Цена: по договоренности"
            )
        except Exception as e:
            print(f"Failed to send to channel {CHANNEL_ID}: {e}")

    db.cart_clear(call.from_user.id)
    await state.clear()

    await call.message.answer(
        TEXT[lang]["order_success"].format(order_id=order_id),
        reply_markup=kb_main(lang, is_admin(call.from_user.id))
    )

    # ✅ Thank you + follow buttons (NEW ORDER)
    await call.message.answer(TEXT[lang]["thanks_new"], reply_markup=kb_follow_links(lang))

    await call.answer()

@dp.callback_query(F.data == "order:cancel")
async def order_cancel(call: CallbackQuery, state: FSMContext):
    await state.clear()
    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"
    await call.message.answer(TEXT[lang]["cancelled"], reply_markup=kb_main(lang, is_admin(call.from_user.id)))
    await call.answer()

# History
@dp.message(F.text.in_(["📜 История", "📜 Buyurtmalar"]))
async def cmd_history(message: Message, state: FSMContext):
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    orders = db.orders_get_user(message.from_user.id)
    if not orders:
        await message.answer(TEXT[lang]["history_empty"], reply_markup=kb_main(lang, is_admin(message.from_user.id)))
        return
    lines = []
    for o in orders[:5]:
        status_icon = {"new": "🆕", "processing": "⚙️", "shipped": "🚚", "delivered": "✅", "cancelled": "❌"}.get(o["status"], "❓")
        lines.append(f"{status_icon} #{o['id']} • {o['created_at'][:10]}")
    await message.answer(TEXT[lang]["history"].format(orders="\n".join(lines)),
                         reply_markup=kb_main(lang, is_admin(message.from_user.id)))

# Admin panel (telegram)
@dp.message(F.text.in_(["🛠 Админ", "🛠 Admin"]))
async def cmd_admin(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"
    await message.answer(TEXT[lang]["admin_menu"], reply_markup=kb_admin(lang))

@dp.callback_query(F.data.startswith("admin:"))
async def admin_action(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        await call.answer("Access denied")
        return

    action = call.data.split(":")[1]
    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"

    if action == "stats":
        stats = db.get_stats_all()
        await call.message.answer(TEXT[lang]["admin_stats"].format(**stats), reply_markup=kb_admin(lang))

    elif action == "new":
        orders = db.orders_get_by_status("new")
        if not orders:
            await call.message.answer("Нет новых заказов" if lang == "ru" else "Yangi buyurtmalar yo'q")
        else:
            for order in orders[:5]:
                items = json.loads(order["items"]) if order.get("items") else []
                items_text = ", ".join([f"{it.get('name','')} x{it.get('qty',1)}" for it in items[:3]])
                text = (
                    f"🆕 Заказ #{order['id']}\n"
                    f"👤 {esc(order['name'])}\n"
                    f"📱 {esc(order['phone'])}\n"
                    f"🏙 {esc(order['city'])}\n"
                    f"🛒 {esc(items_text)}\n"
                    f"💬 Цена: по договоренности"
                )
                await call.message.answer(text, reply_markup=kb_admin_order(order["id"], order["user_id"], lang))

    elif action == "processing":
        orders = db.orders_get_by_status("processing")
        await call.message.answer(
            (f"В обработке: {len(orders)} заказов") if lang == "ru" else (f"Ishlanmoqda: {len(orders)} ta")
        )

    elif action == "export":
        await generate_monthly_report(call.message, lang)

    elif action == "posts":
        await state.set_state(States.admin_post_dow)
        await call.message.answer("Выберите день публикации (Пн–Сб):" if lang == "ru" else "Kun tanlang (Du–Sha):",
                                  reply_markup=kb_dow(lang))

    await call.answer()

# Admin weekly posts
@dp.callback_query(F.data.startswith("dow:"))
async def admin_choose_dow(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    user = db.user_get(call.from_user.id)
    lang = user["lang"] if user else "ru"
    dow = int(call.data.split(":")[1])
    await state.update_data(post_dow=dow)
    await state.set_state(States.admin_post_media)
    await call.message.answer(
        "Теперь отправьте ОДНО сообщение: фото/видео + описание (caption)."
        if lang == "ru" else
        "Endi BITTA xabar yuboring: foto/video + matn (caption)."
    )
    await call.answer()

@dp.message(States.admin_post_media)
async def admin_receive_week_post(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    user = db.user_get(message.from_user.id)
    lang = user["lang"] if user else "ru"

    data = await state.get_data()
    dow = int(data.get("post_dow", 0))
    if dow not in (1, 2, 3, 4, 5, 6):
        await state.clear()
        await message.answer("Сначала выберите день." if lang == "ru" else "Avval kunni tanlang.")
        return

    caption = (message.caption or message.text or "").strip()
    if not caption:
        await message.answer("⚠️ Добавьте описание (текст) к фото/видео." if lang == "ru" else "⚠️ Matn (izoh) qo'shing.")
        return

    media_type = "none"
    file_id = ""
    if message.photo:
        media_type = "photo"
        file_id = message.photo[-1].file_id
    elif message.video:
        media_type = "video"
        file_id = message.video.file_id

    week_key = db.week_key_now(now_tz())
    db.sched_add(dow=dow, media_type=media_type, file_id=file_id, caption=caption, week_key=week_key)
    cnt = db.sched_count_week(week_key)

    await message.answer(
        (f"✅ Добавлено в план недели: <b>{week_key}</b>\n"
         f"📌 День: {dow} (1=Пн ... 6=Сб)\n"
         f"🧾 Постов в этой неделе: <b>{cnt}</b>\n\n"
         "Чтобы добавить ещё — 🛠 Админ → 📰 Посты недели.")
        if lang == "ru" else
        (f"✅ Haftalik reja: <b>{week_key}</b>\n"
         f"📌 Kun: {dow} (1=Du ... 6=Sha)\n"
         f"🧾 Postlar soni: <b>{cnt}</b>\n\n"
         "Yana qo‘shish: 🛠 Admin → 📰 Haftalik postlar.")
    )
    await state.clear()

# Status buttons (admin -> customer notify)
@dp.callback_query(F.data.startswith("order_seen:"))
async def order_seen(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    order_id = int(call.data.split(":")[1])
    db.order_mark_seen(order_id, call.from_user.id)
    await call.answer("✅ Просмотрено")

@dp.callback_query(F.data.startswith("order_process:"))
async def order_process(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    order_id = int(call.data.split(":")[1])
    db.order_update_status(order_id, "processing", call.from_user.id)
    order = db.order_get(order_id)
    if order:
        user_row = db.user_get(order["user_id"])
        lang = user_row["lang"] if user_row else "ru"
        try:
            await bot.send_message(
                order["user_id"],
                (f"⚙️ Заказ #{order_id} в обработке!\nМенеджер скоро свяжется.")
                if lang == "ru" else
                (f"⚙️ Buyurtma #{order_id} ishlanmoqda!\nMenejer tez orada bog'lanadi."),
                reply_markup=kb_main(lang, is_admin(order["user_id"]))
            )
        except Exception as e:
            print(f"Failed to notify user processing: {e}")
    await call.answer("✅ В работе!")

@dp.callback_query(F.data.startswith("order_ship:"))
async def order_ship(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    order_id = int(call.data.split(":")[1])
    db.order_update_status(order_id, "shipped", call.from_user.id)
    order = db.order_get(order_id)
    if order:
        user_row = db.user_get(order["user_id"])
        lang = user_row["lang"] if user_row else "ru"
        try:
            await bot.send_message(
                order["user_id"],
                (f"🚚 Заказ #{order_id} отправлен и уже в пути!\nМы сообщим, когда он будет доставлен.\n⏰ 09:00-21:00")
                if lang == "ru" else
                (f"🚚 Buyurtma #{order_id} jo'natildi va yo'lda!\nYetkazilganda xabar beramiz.\n⏰ 09:00-21:00"),
                reply_markup=kb_main(lang, is_admin(order["user_id"]))
            )
        except Exception as e:
            print(f"Failed to notify user shipped: {e}")
    await call.answer("✅ Отправлен!")

@dp.callback_query(F.data.startswith("order_deliver:"))
async def order_deliver(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    order_id = int(call.data.split(":")[1])
    db.order_update_status(order_id, "delivered", call.from_user.id)
    order = db.order_get(order_id)
    if order:
        user_row = db.user_get(order["user_id"])
        lang = user_row["lang"] if user_row else "ru"
        try:
            await bot.send_message(
                order["user_id"],
                (f"✅ Заказ #{order_id} доставлен!\nСпасибо, что выбрали ZARY & CO 🤍")
                if lang == "ru" else
                (f"✅ Buyurtma #{order_id} yetkazildi!\nZARY & CO ni tanlaganingiz uchun rahmat 🤍"),
                reply_markup=kb_main(lang, is_admin(order["user_id"]))
            )
            # ✅ delivered-thanks
            await bot.send_message(order["user_id"], TEXT[lang]["thanks_delivered"], reply_markup=kb_follow_links(lang))
        except Exception as e:
            print(f"Failed to notify user delivered: {e}")
    await call.answer("✅ Доставлен!")

@dp.callback_query(F.data.startswith("order_cancel:"))
async def order_cancel_admin(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    order_id = int(call.data.split(":")[1])
    db.order_update_status(order_id, "cancelled", call.from_user.id)
    order = db.order_get(order_id)
    if order:
        user_row = db.user_get(order["user_id"])
        lang = user_row["lang"] if user_row else "ru"
        try:
            await bot.send_message(
                order["user_id"],
                (f"❌ Заказ #{order_id} отменён.\nЕсли это ошибка — напишите менеджеру.")
                if lang == "ru" else
                (f"❌ Buyurtma #{order_id} bekor qilindi.\nAgar xato bo‘lsa — menejerga yozing."),
                reply_markup=kb_main(lang, is_admin(order["user_id"]))
            )
        except Exception as e:
            print(f"Failed to notify user cancelled: {e}")
    await call.answer("❌ Отменен!")


# =========================
# REPORTS
# =========================
async def generate_monthly_report(message: Message, lang: str):
    now = now_tz()
    year, month = now.year, now.month

    if db.report_is_sent(year, month):
        await message.answer("Отчет за этот месяц уже отправлен!" if lang == "ru" else "Bu oy hisobot yuborilgan!")
        return

    orders = db.orders_get_monthly(year, month)
    if not orders:
        await message.answer("Нет заказов за этот месяц" if lang == "ru" else "Bu oy buyurtmalar yo'q")
        return

    Path("reports").mkdir(exist_ok=True)
    filename = f"reports/report_{year}_{month:02d}.xlsx"
    total_amount = build_excel_report(filename, orders)

    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, f"📊 Отчет {month:02d}.{year}\n📦 Заказов: {len(orders)}\n💰 Сумма: {total_amount}")
            await bot.send_document(admin_id, FSInputFile(filename))
        except Exception as e:
            print(f"Failed to send report to {admin_id}: {e}")

    db.report_mark_sent(year, month, filename, len(orders), total_amount)
    await message.answer("✅ Отчет отправлен!" if lang == "ru" else "✅ Hisobot yuborildi!")

def build_excel_report(filename: str, orders: List[Dict]) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    headers = ["ID", "Дата", "Клиент", "Телефон", "Город", "Товары", "Сумма", "Статус"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    total_amount = 0
    for order in orders:
        items = json.loads(order["items"]) if order.get("items") else []
        items_str = ", ".join([f"{it.get('name','')} x{it.get('qty',1)}" for it in items])
        amt = int(order.get("total_amount") or 0)

        ws.append([
            order["id"],
            order["created_at"],
            order["name"],
            order["phone"],
            order["city"],
            items_str[:120],
            amt,
            order["status"]
        ])
        total_amount += amt

    wb.save(filename)
    return total_amount

async def cron_send_prev_month_report():
    now = now_tz()
    year, month = prev_month(now)

    if db.report_is_sent(year, month):
        return

    orders = db.orders_get_monthly(year, month)
    if not orders:
        return

    Path("reports").mkdir(exist_ok=True)
    filename = f"reports/report_{year}_{month:02d}.xlsx"
    total_amount = build_excel_report(filename, orders)

    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, f"📊 Автоотчет {month:02d}.{year}\n📦 Заказов: {len(orders)}\n💰 Сумма: {total_amount}")
            await bot.send_document(admin_id, FSInputFile(filename))
        except Exception as e:
            print(f"Auto report failed for {admin_id}: {e}")

    db.report_mark_sent(year, month, filename, len(orders), total_amount)


# =========================
# DAILY CHANNEL POST (Mon–Sat), Sunday reminder
# =========================
async def cron_post_daily_to_channel():
    if not CHANNEL_ID:
        return

    now = now_tz()
    dow = now.isoweekday()  # 1..7

    if dow == 7:
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_message(admin_id, "📌 Воскресенье: загрузите посты на новую неделю (Пн–Сб) → 🛠 Админ → 📰 Посты недели.")
            except Exception:
                pass
        return

    week_key = db.week_key_now(now)
    post = db.sched_get_for_day(dow=dow, week_key=week_key)

    if not post:
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_message(admin_id, f"⚠️ Нет поста на сегодня (день={dow}) для недели {week_key}.")
            except Exception:
                pass
        return

    caption = (post.get("caption") or "").strip() or "🔥 ZARY & CO"
    media_type = post.get("media_type") or "none"
    file_id = post.get("file_id") or ""

    try:
        if media_type == "video" and file_id:
            await bot.send_video(CHANNEL_ID, file_id, caption=caption)
        elif media_type == "photo" and file_id:
            await bot.send_photo(CHANNEL_ID, file_id, caption=caption)
        else:
            await bot.send_message(CHANNEL_ID, caption)

        db.sched_mark_posted(post["id"])
    except Exception as e:
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_message(admin_id, f"❌ Ошибка публикации в канал: {e}")
            except Exception:
                pass


# =========================
# REMINDERS (asyncio loop instead of APScheduler)
# =========================
async def check_reminders():
    orders = db.orders_get_for_reminder()
    if not orders:
        return

    for admin_id in ADMIN_IDS:
        try:
            lines = [f"🆕 #{o['id']} | {esc(o['name'])} | {esc(o['phone'])}" for o in orders[:10]]
            text = "🔔 <b>Напоминание: новые заказы!</b>\n\n" + "\n".join(lines)
            await bot.send_message(admin_id, text)
        except Exception as e:
            print(f"Reminder failed for {admin_id}: {e}")

    for o in orders:
        db.order_update_reminded(o["id"])

async def reminders_loop():
    while True:
        try:
            await check_reminders()
        except Exception as e:
            print("reminders_loop error:", e)
        await asyncio.sleep(30 * 60)  # every 30 minutes


# =========================
# WEB SERVER + CRON + ADMIN PANEL
# =========================
from aiohttp import web

def _month_range_strings(dt: datetime) -> Tuple[str, str]:
    start = dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last = monthrange(dt.year, dt.month)[1]
    end = dt.replace(day=last, hour=23, minute=59, second=59, microsecond=0)
    return start.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S")

def _today_range_strings(dt: datetime) -> Tuple[str, str]:
    start = dt.replace(hour=0, minute=0, second=0, microsecond=0)
    end = dt.replace(hour=23, minute=59, second=59, microsecond=0)
    return start.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S")

def _render_base_html(title: str, body: str) -> str:
    return f"""<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>{esc(title)}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
body{{font-family:Arial, sans-serif; background:#0b0f17; color:#e8eefc; margin:0;}}
a{{color:#7aa2ff; text-decoration:none}}
.header{{padding:16px 20px; border-bottom:1px solid #1c2742; display:flex; justify-content:space-between; align-items:center}}
.container{{padding:20px; max-width:1100px; margin:0 auto}}
.card{{background:#101826; border:1px solid #1c2742; border-radius:14px; padding:16px; margin-bottom:14px}}
.grid{{display:grid; grid-template-columns:repeat(2, 1fr); gap:12px}}
@media (max-width:900px){{.grid{{grid-template-columns:1fr}}}}
.small{{opacity:.85; font-size:13px}}
.badge{{display:inline-block; padding:4px 10px; border-radius:999px; border:1px solid #2b3b66; font-size:12px}}
.table{{width:100%; border-collapse:collapse}}
.table th,.table td{{border-bottom:1px solid #1c2742; padding:10px; text-align:left; font-size:14px}}
.input,select{{background:#0b1220; border:1px solid #1c2742; color:#e8eefc; padding:8px 10px; border-radius:10px}}
.btn{{background:#243a7a; border:none; color:#fff; padding:9px 12px; border-radius:10px; cursor:pointer}}
.btn2{{background:#162542; border:1px solid #2b3b66}}
.row{{display:flex; gap:10px; flex-wrap:wrap; align-items:center}}
</style>
</head>
<body>
<div class="header">
  <div><b>ZARY & CO</b> <span class="badge">Admin</span></div>
  <div class="small">
    <a href="/admin" id="dashLink">Dashboard</a> ·
    <a href="/admin/orders" id="ordersLink">Orders</a>
  </div>
</div>
<div class="container">
{body}
</div>

<script>
(function(){{
  const params = new URLSearchParams(location.search);
  const token = params.get("token") || "";
  const dash = document.getElementById("dashLink");
  const ord = document.getElementById("ordersLink");
  if(dash) dash.href = "/admin?token=" + encodeURIComponent(token);
  if(ord) ord.href = "/admin/orders?token=" + encodeURIComponent(token);
}})();
</script>

</body>
</html>"""

async def health_server():
    app = web.Application()

    async def health(request):
        return web.Response(text="OK", status=200)

    async def cron_monthly(request: web.Request):
        secret = request.query.get("secret", "")
        if not cron_allowed(secret):
            return web.Response(text="Forbidden", status=403)
        await cron_send_prev_month_report()
        return web.Response(text="OK", status=200)

    async def cron_daily(request: web.Request):
        secret = request.query.get("secret", "")
        if not cron_allowed(secret):
            return web.Response(text="Forbidden", status=403)
        await cron_post_daily_to_channel()
        return web.Response(text="OK", status=200)

    # -------- Admin Panel Pages --------
    async def admin_dashboard(request: web.Request):
        token = request.query.get("token", "")
        if not admin_panel_allowed(token):
            return web.Response(text="Forbidden", status=403)

        body = """
        <div class="card">
          <div style="display:flex; justify-content:space-between; align-items:center; gap:10px; flex-wrap:wrap">
            <div>
              <div style="font-size:20px"><b>Аналитика</b></div>
              <div class="small">Данные из SQLite (orders + events). Время: Asia/Tashkent.</div>
            </div>
            <div class="small">Открой: <b>/admin?token=...</b></div>
          </div>
        </div>

        <div class="grid">
          <div class="card">
            <div><b>Сегодня</b></div>
            <div class="small" id="todayStats">Загрузка...</div>
          </div>
          <div class="card">
            <div><b>Этот месяц</b></div>
            <div class="small" id="monthStats">Загрузка...</div>
          </div>
        </div>

        <div class="grid">
          <div class="card">
            <div><b>Топ товары (месяц)</b></div>
            <canvas id="topProducts"></canvas>
          </div>
          <div class="card">
            <div><b>Топ города (месяц)</b></div>
            <canvas id="topCities"></canvas>
          </div>
        </div>

        <div class="grid">
          <div class="card">
            <div><b>RU vs UZ (месяц)</b></div>
            <canvas id="langChart"></canvas>
          </div>
          <div class="card">
            <div><b>Конверсия (месяц)</b></div>
            <div class="small" id="funnelBox">Загрузка...</div>
          </div>
        </div>

        <script>
        async function api(path){
          const params = new URLSearchParams(location.search);
          const token = params.get("token") || "";
          const r = await fetch(path + "?token=" + encodeURIComponent(token));
          if(!r.ok) throw new Error("API error");
          return await r.json();
        }

        function formatStats(s){
          return `Всего: ${s.total} | Новые: ${s.new} | В обработке: ${s.processing} | Отправлено: ${s.shipped} | Доставлено: ${s.delivered} | Отменено: ${s.cancelled}`;
        }

        (async () => {
          const today = await api("/admin/api/stats/today");
          const month = await api("/admin/api/stats/month");

          document.getElementById("todayStats").textContent = formatStats(today.stats);
          document.getElementById("monthStats").textContent = formatStats(month.stats);

          const tp = month.top_products;
          new Chart(document.getElementById("topProducts"), {
            type: "bar",
            data: {
              labels: tp.map(x => x.name),
              datasets: [{ label: "Кол-во", data: tp.map(x => x.count) }]
            },
            options: {
              plugins: { legend: { display: false } }
            }
          });

          const tc = month.top_cities;
          new Chart(document.getElementById("topCities"), {
            type: "bar",
            data: {
              labels: tc.map(x => x.city),
              datasets: [{ label: "Заказы", data: tc.map(x => x.count) }]
            },
            options: {
              plugins: { legend: { display: false } }
            }
          });

          const lc = month.ru_vs_uz;
          new Chart(document.getElementById("langChart"), {
            type: "doughnut",
            data: {
              labels: ["RU", "UZ", "Unknown"],
              datasets: [{ data: [lc.ru, lc.uz, lc.unknown] }]
            }
          });

          const f = month.funnel;
          document.getElementById("funnelBox").innerHTML =
            `<div class="small">Добавили в корзину: <b>${f.cart_add}</b></div>` +
            `<div class="small">Оформили заказ: <b>${f.order_created}</b></div>` +
            `<div class="small">Конверсия: <b>${f.conversion}%</b></div>`;
        })().catch(e => {
          document.getElementById("todayStats").textContent = "Ошибка загрузки";
          document.getElementById("monthStats").textContent = "Ошибка загрузки";
          document.getElementById("funnelBox").textContent = "Ошибка загрузки";
        });
        </script>
        """
        return web.Response(text=_render_base_html("ZARY Admin Dashboard", body), content_type="text/html")

    async def admin_orders(request: web.Request):
        token = request.query.get("token", "")
        if not admin_panel_allowed(token):
            return web.Response(text="Forbidden", status=403)

        body = """
        <div class="card">
          <div style="display:flex; justify-content:space-between; align-items:center; gap:10px; flex-wrap:wrap">
            <div>
              <div style="font-size:20px"><b>Заказы</b></div>
              <div class="small">Фильтр: статус/город/телефон. Можно менять статус прямо тут.</div>
            </div>
          </div>

          <div class="row" style="margin-top:12px">
            <select class="input" id="st">
              <option value="">Все статусы</option>
              <option value="new">new</option>
              <option value="processing">processing</option>
              <option value="shipped">shipped</option>
              <option value="delivered">delivered</option>
              <option value="cancelled">cancelled</option>
            </select>
            <input class="input" id="city" placeholder="Город (часть)"/>
            <input class="input" id="phone" placeholder="Телефон (часть)"/>
            <button class="btn" id="go">Показать</button>
          </div>
        </div>

        <div class="card">
          <div id="tableWrap" class="small">Загрузка...</div>
        </div>

        <script>
        function qs(id){ return document.getElementById(id); }
        function token(){ return new URLSearchParams(location.search).get("token") || ""; }

        async function api(path, opts){
          const glue = path.includes("?") ? "&" : "?";
          const r = await fetch(path + glue + "token=" + encodeURIComponent(token()), opts || {});
          if(!r.ok) throw new Error("API error");
          return await r.json();
        }

        function statusSelect(id, current){
          const st = ["new","processing","shipped","delivered","cancelled"];
          return `<select class="input" data-oid="${id}">
            ${st.map(s => `<option value="${s}" ${s===current?"selected":""}>${s}</option>`).join("")}
          </select>`;
        }

        function render(rows){
          if(!rows.length){
            qs("tableWrap").innerHTML = "Нет заказов по фильтру";
            return;
          }
          const html = `
          <table class="table">
            <thead>
              <tr>
                <th>ID</th><th>Дата</th><th>Клиент</th><th>Телефон</th><th>Город</th><th>Товары</th><th>Статус</th><th>Действие</th>
              </tr>
            </thead>
            <tbody>
              ${rows.map(o => `
                <tr>
                  <td>#${o.id}</td>
                  <td>${(o.created_at||"").slice(0,16)}</td>
                  <td>${o.name||""}</td>
                  <td>${o.phone||""}</td>
                  <td>${o.city||""}</td>
                  <td>${(o.items_preview||"")}</td>
                  <td>${statusSelect(o.id, o.status)}</td>
                  <td><button class="btn btn2" data-save="${o.id}">Сохранить</button></td>
                </tr>
              `).join("")}
            </tbody>
          </table>`;
          qs("tableWrap").innerHTML = html;

          document.querySelectorAll("[data-save]").forEach(btn => {
            btn.addEventListener("click", async () => {
              const id = btn.getAttribute("data-save");
              const sel = document.querySelector(`select[data-oid="${id}"]`);
              const status = sel.value;
              btn.textContent = "...";
              try{
                await api("/admin/api/order/status", {
                  method:"POST",
                  headers: {"Content-Type":"application/json"},
                  body: JSON.stringify({order_id: Number(id), status})
                });
                btn.textContent = "OK";
                setTimeout(()=>btn.textContent="Сохранить", 800);
              }catch(e){
                btn.textContent = "ERR";
                setTimeout(()=>btn.textContent="Сохранить", 1200);
              }
            });
          });
        }

        async function load(){
          qs("tableWrap").textContent = "Загрузка...";
          const st = qs("st").value;
          const city = qs("city").value.trim();
          const phone = qs("phone").value.trim();
          const path = `/admin/api/orders?status=${encodeURIComponent(st)}&city=${encodeURIComponent(city)}&phone=${encodeURIComponent(phone)}`;
          const data = await api(path);
          render(data.orders);
        }

        qs("go").addEventListener("click", load);
        load();
        </script>
        """
        return web.Response(text=_render_base_html("ZARY Admin Orders", body), content_type="text/html")

    # -------- Admin Panel APIs --------
    async def api_stats_today(request: web.Request):
        token = request.query.get("token", "")
        if not admin_panel_allowed(token):
            return web.json_response({"error": "forbidden"}, status=403)
        dt = now_tz()
        start, end = _today_range_strings(dt)
        stats = db.stats_range(start, end)
        return web.json_response({"range": {"start": start, "end": end}, "stats": stats})

    async def api_stats_month(request: web.Request):
        token = request.query.get("token", "")
        if not admin_panel_allowed(token):
            return web.json_response({"error": "forbidden"}, status=403)
        dt = now_tz()
        start, end = _month_range_strings(dt)
        stats = db.stats_range(start, end)
        top_products = [{"name": n, "count": c} for n, c in db.top_products_range(start, end)]
        top_cities = [{"city": n, "count": c} for n, c in db.top_cities_range(start, end)]
        ruuz = db.ru_vs_uz_range(start, end)
        funnel = db.funnel_range(start, end)
        return web.json_response({
            "range": {"start": start, "end": end},
            "stats": stats,
            "top_products": top_products,
            "top_cities": top_cities,
            "ru_vs_uz": ruuz,
            "funnel": funnel
        })

    async def api_orders(request: web.Request):
        token = request.query.get("token", "")
        if not admin_panel_allowed(token):
            return web.json_response({"error": "forbidden"}, status=403)

        status = request.query.get("status", "").strip()
        city = request.query.get("city", "").strip()
        phone = request.query.get("phone", "").strip()

        rows = db.orders_filter(status=status, city=city, phone_q=phone, limit=200)
        orders = []
        for o in rows:
            try:
                items = json.loads(o.get("items") or "[]")
                items_preview = ", ".join([f"{it.get('name','')} x{it.get('qty',1)}" for it in items[:3]])
            except Exception:
                items_preview = ""
            orders.append({
                "id": o["id"],
                "created_at": o.get("created_at",""),
                "name": o.get("name",""),
                "phone": o.get("phone",""),
                "city": o.get("city",""),
                "status": o.get("status",""),
                "items_preview": items_preview
            })
        return web.json_response({"orders": orders})

    async def api_order_status(request: web.Request):
        token = request.query.get("token", "")
        if not admin_panel_allowed(token):
            return web.json_response({"error": "forbidden"}, status=403)

        try:
            payload = await request.json()
        except Exception:
            return web.json_response({"error": "bad_json"}, status=400)

        order_id = int(payload.get("order_id", 0) or 0)
        status = (payload.get("status", "") or "").strip()
        if order_id <= 0 or status not in ("new","processing","shipped","delivered","cancelled"):
            return web.json_response({"error": "bad_params"}, status=400)

        db.order_update_status(order_id, status, PRIMARY_ADMIN)

        # notify customer
        order = db.order_get(order_id)
        if order:
            user_row = db.user_get(order["user_id"])
            lang = user_row["lang"] if user_row else "ru"
            try:
                if status == "processing":
                    await bot.send_message(
                        order["user_id"],
                        (f"⚙️ Заказ #{order_id} в обработке!\nМенеджер скоро свяжется.")
                        if lang == "ru" else
                        (f"⚙️ Buyurtma #{order_id} ishlanmoqda!\nMenejer tez orada bog'lanadi."),
                        reply_markup=kb_main(lang, is_admin(order["user_id"]))
                    )
                elif status == "shipped":
                    await bot.send_message(
                        order["user_id"],
                        (f"🚚 Заказ #{order_id} отправлен и уже в пути!\nМы сообщим, когда он будет доставлен.\n⏰ 09:00-21:00")
                        if lang == "ru" else
                        (f"🚚 Buyurtma #{order_id} jo'natildi va yo'lda!\nYetkazilganda xabar beramiz.\n⏰ 09:00-21:00"),
                        reply_markup=kb_main(lang, is_admin(order["user_id"]))
                    )
                elif status == "delivered":
                    await bot.send_message(
                        order["user_id"],
                        (f"✅ Заказ #{order_id} доставлен!\nСпасибо, что выбрали ZARY & CO 🤍")
                        if lang == "ru" else
                        (f"✅ Buyurtma #{order_id} yetkazildi!\nZARY & CO ni tanlaganingiz uchun rahmat 🤍"),
                        reply_markup=kb_main(lang, is_admin(order["user_id"]))
                    )
                    await bot.send_message(order["user_id"], TEXT[lang]["thanks_delivered"], reply_markup=kb_follow_links(lang))
                elif status == "cancelled":
                    await bot.send_message(
                        order["user_id"],
                        (f"❌ Заказ #{order_id} отменён.\nЕсли это ошибка — напишите менеджеру.")
                        if lang == "ru" else
                        (f"❌ Buyurtma #{order_id} bekor qilindi.\nAgar xato bo‘lsa — menejerga yozing."),
                        reply_markup=kb_main(lang, is_admin(order["user_id"]))
                    )
            except Exception as e:
                print("Web notify failed:", e)

        return web.json_response({"ok": True})

    # routes
    app.router.add_get("/", health)
    app.router.add_get("/health", health)
    app.router.add_get("/cron/monthly", cron_monthly)
    app.router.add_get("/cron/daily", cron_daily)

    app.router.add_get("/admin", admin_dashboard)
    app.router.add_get("/admin/orders", admin_orders)

    app.router.add_get("/admin/api/stats/today", api_stats_today)
    app.router.add_get("/admin/api/stats/month", api_stats_month)
    app.router.add_get("/admin/api/orders", api_orders)
    app.router.add_post("/admin/api/order/status", api_order_status)

    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", PORT)
    await site.start()
    print(f"✅ Health/Admin server on port {PORT}")


# =========================
# MAIN
# =========================
async def main():
    await health_server()
    asyncio.create_task(reminders_loop())
    print(f"✅ Bot started with {len(ADMIN_IDS)} admins: {ADMIN_IDS}")
    if CHANNEL_ID:
        print(f"✅ Channel enabled: {CHANNEL_ID}")
    if CRON_SECRET:
        print("✅ Cron endpoints enabled: /cron/monthly /cron/daily")
    print("✅ Admin panel: /admin?token=YOUR_TOKEN")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())

